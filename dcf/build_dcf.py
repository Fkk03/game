#!/usr/bin/env python3
"""Build dcf_models.xlsx — 5-year DCF models for 9 tech companies.

Every projection cell is a live Excel formula driven by labeled input cells
(blue font = hardcoded input, yellow fill = key levers to play with), so the
workbook recalculates when assumptions change. Run the xlsx skill's
scripts/recalc.py after generating (openpyxl writes formulas without cached
values).

Methodology (same engine for every company):
  Revenue_t = Revenue_{t-1} * (1 + g_t)
  EBIT_t = Revenue_t * margin_t          (economic margin — treats SBC as a real cost)
  NOPAT_t = EBIT_t * (1 - tax)
  uFCF_t = NOPAT_t + D&A_t - Capex_t - dNWC_t
  PV at WACC with mid-year convention.
  Terminal value: Gordon growth on a NORMALIZED terminal uFCF margin applied to
  year-5 revenue (decouples the perpetuity from year-5 transition/cycle-state
  margins — mid-cycle for memory names, steady-state for high-SBC software).
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

VALUATION_DATE = "August 24, 2026"

# ---------------------------------------------------------------- styling ---
F = "Arial"
TITLE = Font(name=F, size=14, bold=True, color="1F3864")
SUB = Font(name=F, size=9, italic=True, color="595959")
HDR = Font(name=F, size=10, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="1F3864")
COLHDR = Font(name=F, size=10, bold=True)
COLHDR_FILL = PatternFill("solid", fgColor="D9E2F3")
LBL = Font(name=F, size=10)
LBL_B = Font(name=F, size=10, bold=True)
INPUT = Font(name=F, size=10, color="0000FF")          # blue = hardcoded input
FORM = Font(name=F, size=10)                            # black = formula
FORM_B = Font(name=F, size=10, bold=True)
LINK = Font(name=F, size=10, color="008000")            # green = cross-sheet link
NOTE = Font(name=F, size=9, italic=True, color="808080")
YELLOW = PatternFill("solid", fgColor="FFFF00")         # key levers
GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")
SENS_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOPLINE = Border(top=Side(style="thin", color="404040"))

MONEY_USD = '"$"#,##0;("$"#,##0);"-"'
MONEY_PLAIN = '#,##0;(#,##0);"-"'
PRICE_USD = '"$"#,##0.00'
PRICE_PLAIN = '#,##0'
PCT1 = '0.0%;(0.0%)'
SHARES_FMT = '#,##0.0'
MULT = '0.0"x"'
NUM2 = '0.00'

COLS = ["C", "D", "E", "F", "G"]

# row map for company sheets
R_REV, R_G, R_M, R_EBIT, R_TAX, R_NOPAT = 15, 16, 17, 18, 19, 20
R_DAP, R_DA, R_CXP, R_CX, R_NWCP, R_NWC = 21, 22, 23, 24, 25, 26
R_FCF, R_FCFM, R_PER, R_DF, R_PV = 27, 28, 30, 31, 32
R_TVH, R_TM, R_TFCF, R_TV, R_PVTV, R_TMULT = 34, 35, 36, 37, 38, 39
R_VALH, R_PVX, R_PVT, R_EV, R_TVPCT, R_NC, R_EQ, R_VPS, R_PX, R_UP = (
    41, 42, 43, 44, 45, 46, 47, 48, 49, 50)
R_SENSH, R_SENSHDR = 52, 53
R_NOTESH = 60


def style_cell(ws, addr, value, font=FORM, fmt=None, fill=None, align=None, border=None):
    c = ws[addr]
    c.value = value
    c.font = font
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if align:
        c.alignment = align
    if border:
        c.border = border
    return c


def header_band(ws, row, text):
    style_cell(ws, f"A{row}", text, HDR, fill=HDR_FILL)
    for col in "BCDEFG":
        ws[f"{col}{row}"].fill = HDR_FILL


def python_dcf(co):
    """Numeric mirror of the sheet formulas — used only to write sanity notes."""
    rev, out = co["base_revenue"], []
    for i in range(5):
        rev_next = rev * (1 + co["growth"][i])
        ebit = rev_next * co["ebit_margin"][i]
        nopat = ebit * (1 - co["tax"])
        fcf = (nopat + rev_next * co["da_pct"][i] - rev_next * co["capex_pct"][i]
               - (rev_next - rev) * co["nwc_pct"])
        out.append((rev_next, fcf))
        rev = rev_next
    w, g = co["wacc"], co["tgr"]
    pv = sum(f / (1 + w) ** (i + 0.5) for i, (_, f) in enumerate(out))
    tfcf = out[-1][0] * (1 + g) * co["term_margin"]
    tv = tfcf / (w - g)
    ev = pv + tv / (1 + w) ** 4.5
    eq = ev + co["net_cash"]
    vps = eq * co["ps_factor"] / co["shares"]
    return dict(fy5_rev=out[-1][0], fy5_fcf=out[-1][1], ev=ev, vps=vps)


def build_company_sheet(wb, co):
    ws = wb.create_sheet(co["ticker"])
    money = MONEY_USD if co["currency"] == "US$" else MONEY_PLAIN
    price_fmt = PRICE_USD if co["currency"] == "US$" else PRICE_PLAIN
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14
    for col in COLS:
        ws.column_dimensions[col].width = 12
    ws.sheet_view.showGridLines = False

    # ---- header -------------------------------------------------------------
    style_cell(ws, "A1", f'{co["name"]} ({co["exchange"]}: {co["ticker_disp"]}) — 5-Year DCF', TITLE)
    style_cell(ws, "A2", f'Valuation date: {VALUATION_DATE}  |  All figures in {co["unit"]} unless noted  |  {co["fye_note"]}', SUB)

    # ---- key inputs ---------------------------------------------------------
    header_band(ws, 4, "KEY INPUTS  (blue = hardcoded input; yellow = key levers)")
    inputs = [
        ("A5", f'Current share price ({co["currency"]})', "B5", co["price"], price_fmt, co["price_note"]),
        ("A6", "Diluted shares outstanding (mm)", "B6", co["shares"], SHARES_FMT, co["shares_note"]),
        ("A7", f'Net cash / (net debt), {co["unit_short"]}', "B7", co["net_cash"], money, co["net_cash_note"]),
        ("A8", "Tax rate on EBIT", "B8", co["tax"], PCT1, co.get("tax_note", "")),
        ("A9", "WACC (discount rate)", "B9", co["wacc"], PCT1, co.get("wacc_note", "")),
        ("A10", "Terminal growth rate", "B10", co["tgr"], PCT1, co.get("tgr_note", "")),
        ("A11", f'Per-share unit factor ({co["unit_short"]} → {co["currency"]}/sh)', "B11", co["ps_factor"], "#,##0", "1 for $mm & mm shares; 1,000 for ₩bn/¥bn & mm shares"),
    ]
    for laddr, label, vaddr, val, fmt, note in inputs:
        style_cell(ws, laddr, label, LBL)
        fill = YELLOW if vaddr in ("B9", "B10") else None
        style_cell(ws, vaddr, val, INPUT, fmt=fmt, fill=fill, border=BOX)
        if note:
            style_cell(ws, f"C{laddr[1:]}", note, NOTE)

    # ---- projection table ---------------------------------------------------
    header_band(ws, 13, "PROJECTIONS")
    style_cell(ws, "A14", co["unit_short"], COLHDR, fill=COLHDR_FILL)
    style_cell(ws, "B14", co["base_label"], COLHDR, fill=COLHDR_FILL,
               align=Alignment(horizontal="center"))
    for i, col in enumerate(COLS):
        style_cell(ws, f"{col}14", co["years"][i], COLHDR, fill=COLHDR_FILL,
                   align=Alignment(horizontal="center"))

    style_cell(ws, f"A{R_REV}", "Revenue", LBL_B)
    style_cell(ws, f"B{R_REV}", co["base_revenue"], INPUT, fmt=money, border=BOX)
    style_cell(ws, f"A{R_G}", "  Revenue growth %", LBL)
    style_cell(ws, f"B{R_G}", co["base_growth"], NOTE, fmt=PCT1)
    style_cell(ws, f"A{R_M}", "  EBIT margin % (economic, incl. SBC)", LBL)
    style_cell(ws, f"B{R_M}", co["base_margin"], NOTE, fmt=PCT1)
    style_cell(ws, f"A{R_EBIT}", "EBIT", LBL)
    style_cell(ws, f"B{R_EBIT}", f"=B{R_REV}*B{R_M}", NOTE, fmt=money)
    style_cell(ws, f"A{R_TAX}", "Cash taxes on EBIT", LBL)
    style_cell(ws, f"A{R_NOPAT}", "NOPAT", LBL)
    style_cell(ws, f"A{R_DAP}", "  D&A % of revenue", LBL)
    style_cell(ws, f"A{R_DA}", "(+) D&A", LBL)
    style_cell(ws, f"A{R_CXP}", "  Capex % of revenue", LBL)
    style_cell(ws, f"A{R_CX}", "(–) Capex", LBL)
    style_cell(ws, f"A{R_NWCP}", "  ΔNWC as % of Δrevenue (neg = cash source)", LBL)
    style_cell(ws, f"A{R_NWC}", "(–) Investment in working capital", LBL)
    style_cell(ws, f"A{R_FCF}", "Unlevered free cash flow", LBL_B)
    style_cell(ws, f"A{R_FCFM}", "  uFCF margin %", LBL)
    style_cell(ws, f"A{R_PER}", "Discount period (mid-year conv.)", LBL)
    style_cell(ws, f"A{R_DF}", "Discount factor", LBL)
    style_cell(ws, f"A{R_PV}", "PV of uFCF", LBL_B)

    for i, col in enumerate(COLS):
        prev = "B" if i == 0 else COLS[i - 1]
        style_cell(ws, f"{col}{R_G}", co["growth"][i], INPUT, fmt=PCT1, fill=YELLOW, border=BOX)
        style_cell(ws, f"{col}{R_M}", co["ebit_margin"][i], INPUT, fmt=PCT1, fill=YELLOW, border=BOX)
        style_cell(ws, f"{col}{R_REV}", f"={prev}{R_REV}*(1+{col}{R_G})", FORM_B, fmt=money)
        style_cell(ws, f"{col}{R_EBIT}", f"={col}{R_REV}*{col}{R_M}", FORM, fmt=money)
        style_cell(ws, f"{col}{R_TAX}", f"=-{col}{R_EBIT}*$B$8", FORM, fmt=money)
        style_cell(ws, f"{col}{R_NOPAT}", f"={col}{R_EBIT}+{col}{R_TAX}", FORM, fmt=money)
        style_cell(ws, f"{col}{R_DAP}", co["da_pct"][i], INPUT, fmt=PCT1, border=BOX)
        style_cell(ws, f"{col}{R_DA}", f"={col}{R_REV}*{col}{R_DAP}", FORM, fmt=money)
        style_cell(ws, f"{col}{R_CXP}", co["capex_pct"][i], INPUT, fmt=PCT1, border=BOX)
        style_cell(ws, f"{col}{R_CX}", f"=-{col}{R_REV}*{col}{R_CXP}", FORM, fmt=money)
        style_cell(ws, f"{col}{R_NWCP}", co["nwc_pct"], INPUT, fmt=PCT1, border=BOX)
        style_cell(ws, f"{col}{R_NWC}", f"=-({col}{R_REV}-{prev}{R_REV})*{col}{R_NWCP}", FORM, fmt=money)
        style_cell(ws, f"{col}{R_FCF}",
                   f"={col}{R_NOPAT}+{col}{R_DA}+{col}{R_CX}+{col}{R_NWC}",
                   FORM_B, fmt=money, border=TOPLINE)
        style_cell(ws, f"{col}{R_FCFM}", f"={col}{R_FCF}/{col}{R_REV}", FORM, fmt=PCT1)
        if i == 0:
            style_cell(ws, f"{col}{R_PER}", 0.5, INPUT, fmt=NUM2)
        else:
            style_cell(ws, f"{col}{R_PER}", f"={prev}{R_PER}+1", FORM, fmt=NUM2)
        style_cell(ws, f"{col}{R_DF}", f"=1/POWER(1+$B$9,{col}{R_PER})", FORM, fmt="0.000")
        style_cell(ws, f"{col}{R_PV}", f"={col}{R_FCF}*{col}{R_DF}", FORM_B, fmt=money)

    # ---- terminal value -----------------------------------------------------
    header_band(ws, R_TVH, "TERMINAL VALUE (Gordon growth on normalized margin)")
    style_cell(ws, f"A{R_TM}", "Normalized terminal uFCF margin (× yr-5 revenue)", LBL)
    style_cell(ws, f"B{R_TM}", co["term_margin"], INPUT, fmt=PCT1, fill=YELLOW, border=BOX)
    style_cell(ws, f"C{R_TM}", co.get("term_margin_note", ""), NOTE)
    style_cell(ws, f"A{R_TFCF}", "Terminal uFCF (yr-5 revenue × (1+g) × margin)", LBL)
    style_cell(ws, f"B{R_TFCF}", f"=G{R_REV}*(1+$B$10)*B{R_TM}", FORM, fmt=money)
    style_cell(ws, f"A{R_TV}", "Terminal value at end of year 5", LBL)
    style_cell(ws, f"B{R_TV}", f"=B{R_TFCF}/($B$9-$B$10)", FORM, fmt=money)
    style_cell(ws, f"A{R_PVTV}", "PV of terminal value", LBL)
    style_cell(ws, f"B{R_PVTV}", f"=B{R_TV}*G{R_DF}", FORM, fmt=money)
    style_cell(ws, f"A{R_TMULT}", "Implied terminal EV / yr-5 uFCF", LBL)
    style_cell(ws, f"B{R_TMULT}", f"=B{R_TV}/G{R_FCF}", FORM, fmt=MULT)

    # ---- valuation ----------------------------------------------------------
    header_band(ws, R_VALH, "VALUATION")
    val_rows = [
        (R_PVX, "PV of explicit-period uFCF (yrs 1–5)", f"=SUM(C{R_PV}:G{R_PV})", money, FORM),
        (R_PVT, "PV of terminal value", f"=B{R_PVTV}", money, FORM),
        (R_EV, "Enterprise value", f"=B{R_PVX}+B{R_PVT}", money, FORM_B),
        (R_TVPCT, "  Terminal value % of EV", f"=B{R_PVT}/B{R_EV}", PCT1, FORM),
        (R_NC, "(+) Net cash / (net debt)", "=B7", money, FORM),
        (R_EQ, "Equity value", f"=B{R_EV}+B{R_NC}", money, FORM_B),
        (R_VPS, f'Implied value per share ({co["currency"]})', f"=B{R_EQ}*$B$11/B6", price_fmt, FORM_B),
        (R_PX, f'Current share price ({co["currency"]})', "=B5", price_fmt, FORM),
        (R_UP, "Implied upside / (downside)", f"=B{R_VPS}/B{R_PX}-1", PCT1, FORM_B),
    ]
    for row, label, formula, fmt, font in val_rows:
        style_cell(ws, f"A{row}", label, LBL_B if font is FORM_B else LBL)
        style_cell(ws, f"B{row}", formula, font, fmt=fmt)
    ws[f"B{R_VPS}"].fill = GREEN_FILL
    ws[f"B{R_UP}"].fill = GREEN_FILL

    # ---- sensitivity --------------------------------------------------------
    header_band(ws, R_SENSH, "SENSITIVITY — implied value per share (WACC ↓ × terminal growth →)")
    style_cell(ws, f"B{R_SENSHDR}", "WACC \\ g", COLHDR, fill=COLHDR_FILL,
               align=Alignment(horizontal="center"))
    tg_offsets = [-0.010, -0.005, 0.0, 0.005, 0.010]
    w_offsets = [-0.010, -0.005, 0.0, 0.005, 0.010]
    for j, col in enumerate(COLS):
        style_cell(ws, f"{col}{R_SENSHDR}",
                   f"=$B$10+{tg_offsets[j]}" if tg_offsets[j] else "=$B$10",
                   FORM, fmt=PCT1, fill=COLHDR_FILL, align=Alignment(horizontal="center"))
    for i2, off in enumerate(w_offsets):
        row = R_SENSHDR + 1 + i2
        style_cell(ws, f"B{row}", f"=$B$9+{off}" if off else "=$B$9", FORM, fmt=PCT1, fill=COLHDR_FILL)
        for j, col in enumerate(COLS):
            w_ref = f"$B{row}"
            g_ref = f"{col}${R_SENSHDR}"
            formula = (
                f"=(SUMPRODUCT($C${R_FCF}:$G${R_FCF},1/POWER(1+{w_ref},$C${R_PER}:$G${R_PER}))"
                f"+$G${R_REV}*(1+{g_ref})*$B${R_TM}/({w_ref}-{g_ref})/POWER(1+{w_ref},$G${R_PER})"
                f"+$B$7)*$B$11/$B$6"
            )
            fill = SENS_FILL if (off == 0 and tg_offsets[j] == 0) else None
            style_cell(ws, f"{col}{row}", formula, FORM, fmt=price_fmt, fill=fill, border=BOX)

    # ---- notes --------------------------------------------------------------
    header_band(ws, R_NOTESH, "ASSUMPTION NOTES & SOURCES")
    stats = python_dcf(co)
    mcap = co["price"] * co["shares"] / co["ps_factor"]
    implied_mult = (mcap - co["net_cash"]) / stats["fy5_fcf"] if stats["fy5_fcf"] > 0 else None
    auto = (f"Cross-check: market cap ≈ {mcap:,.0f} {co['unit_short']}. At the current price, EV ≈ "
            f"{mcap - co['net_cash']:,.0f} = {implied_mult:,.1f}x this model's year-5 economic uFCF "
            f"({stats['fy5_fcf']:,.0f}) — vs the model's implied terminal multiple shown above. "
            "The gap between those two numbers is what the market is paying for growth beyond year 5."
            ) if implied_mult else ""
    nrow = R_NOTESH + 1
    for note in co["notes"] + ([auto] if auto else []) + ["Sources: " + "; ".join(co["sources"])]:
        ws.merge_cells(f"A{nrow}:G{nrow}")
        style_cell(ws, f"A{nrow}", note, NOTE, align=Alignment(wrap_text=True, vertical="top"))
        ws.row_dimensions[nrow].height = max(12, 12 * (1 + len(note) // 105))
        nrow += 1

    ws.freeze_panes = "B15"
    return ws


def build_summary(wb, companies):
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False
    widths = {"A": 8, "B": 26, "C": 8, "D": 12, "E": 13, "F": 12, "G": 12, "H": 12,
              "I": 9, "J": 9, "K": 10, "L": 10, "M": 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    style_cell(ws, "A1", "5-Year DCF Models — AI Infrastructure & Data Platforms", TITLE)
    style_cell(ws, "A2", f"Valuation date: {VALUATION_DATE}. Prices = latest close/quote per tab. "
                          "Green cells link live to each company tab; edit assumptions there (yellow cells).", SUB)
    headers = ["Ticker", "Company", "Ccy", "Price", "DCF value/sh", "Upside/(dn)",
               "Rev CAGR y1–5", "Yr-5 EBIT mgn", "WACC", "Term. g", "Term. mgn", "TV % of EV", "Fiscal window"]
    for j, h in enumerate(headers):
        c = get_column_letter(j + 1)
        style_cell(ws, f"{c}4", h, HDR, fill=HDR_FILL,
                   align=Alignment(horizontal="center", wrap_text=True))
    row = 5
    for co in companies:
        t = co["ticker"]
        price_fmt = PRICE_USD if co["currency"] == "US$" else PRICE_PLAIN
        style_cell(ws, f"A{row}", co["ticker_disp"], LBL_B)
        style_cell(ws, f"B{row}", co["name"], LBL)
        style_cell(ws, f"C{row}", co["currency"], LBL, align=Alignment(horizontal="center"))
        style_cell(ws, f"D{row}", f"={t}!B5", LINK, fmt=price_fmt)
        style_cell(ws, f"E{row}", f"={t}!B{R_VPS}", LINK, fmt=price_fmt)
        style_cell(ws, f"F{row}", f"={t}!B{R_UP}", LINK, fmt=PCT1)
        style_cell(ws, f"G{row}", f"=POWER({t}!G{R_REV}/{t}!B{R_REV},1/5)-1", LINK, fmt=PCT1)
        style_cell(ws, f"H{row}", f"={t}!G{R_M}", LINK, fmt=PCT1)
        style_cell(ws, f"I{row}", f"={t}!B9", LINK, fmt=PCT1)
        style_cell(ws, f"J{row}", f"={t}!B10", LINK, fmt=PCT1)
        style_cell(ws, f"K{row}", f"={t}!B{R_TM}", LINK, fmt=PCT1)
        style_cell(ws, f"L{row}", f"={t}!B{R_TVPCT}", LINK, fmt=PCT1)
        style_cell(ws, f"M{row}", co["base_label"] + " → " + co["years"][-1], NOTE)
        for c2 in "ABCDEFGHIJKLM":
            ws[f"{c2}{row}"].border = BOX
        row += 1
    style_cell(ws, f"A{row + 1}",
               "Economic EBIT margins treat stock-based compensation as a real expense (between GAAP and non-GAAP "
               "operating margin) — see ReadMe. KRW/JPY names valued in local currency.", NOTE)
    ws.merge_cells(f"A{row + 1}:M{row + 1}")
    return ws


def build_readme(wb):
    ws = wb.create_sheet("ReadMe")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 120
    lines = [
        ("5-Year DCF Models — Methodology & How to Use", TITLE),
        (f"Built {VALUATION_DATE}. Estimates anchored to company guidance and analyst-day / investor-day targets; "
         "all base figures from latest filings and earnings releases (sources on each tab).", SUB),
        ("", LBL),
        ("HOW TO USE", LBL_B),
        ("• Blue cells are hardcoded inputs; yellow-filled cells are the key levers (revenue growth, EBIT margin, "
         "WACC, terminal growth, terminal margin). Edit those — everything else recalculates.", LBL),
        ("• Each company tab: inputs → 5-yr projection → terminal value → EV bridge → per-share value → "
         "WACC × terminal-growth sensitivity grid → assumption notes with sources.", LBL),
        ("• The Summary tab links live to every company tab (green cells).", LBL),
        ("", LBL),
        ("METHODOLOGY", LBL_B),
        ("• Unlevered DCF: uFCF = EBIT × (1 − tax) + D&A − capex − ΔNWC. EBIT margins are 'economic' margins that "
         "treat stock-based compensation as a real expense (they sit between GAAP and non-GAAP operating margins). "
         "This avoids the classic software-DCF trap of capitalizing non-GAAP FCF while ignoring SBC dilution — it is "
         "the main reason the high-SBC software names (SNOW ~29% SBC/revenue, NTNX ~13%, Everpure ~13%) show DCF "
         "values far below market. To see a street-style non-GAAP valuation instead, set the yellow margin cells to "
         "the non-GAAP operating margins quoted in each tab's notes.", LBL),
        ("• Mid-year discounting convention (periods 0.5 … 4.5). For companies whose fiscal year 1 is already partly "
         "elapsed (Jan/Feb/Jun/Jul FYE), this slightly flatters PV; immaterial to conclusions.", LBL),
        ("• Terminal value: Gordon growth on a NORMALIZED terminal uFCF margin × year-5 revenue. For the cyclical "
         "memory names (SK hynix, Samsung, Sandisk, Kioxia) the terminal margin is set near MID-CYCLE, not the "
         "current super-cycle peak, so the perpetuity does not capitalize peak earnings. For the software names it "
         "is set at estimated steady-state economics rather than the year-5 point on the transition path. Each margin "
         "reconciles as: steady-state economic EBIT × (1−tax) + D&A − capex − growth×NWC (the derivation is on each "
         "tab). Discounting is mid-year-consistent: PV(TV) = [TFCF/(WACC−g)] × DF(4.5) is the exact value of a "
         "perpetuity whose payments land at mid-year, matching the explicit period.", LBL),
        ("• Net cash = latest-quarter cash + investments − total debt (incl. converts). Diluted shares = latest "
         "reported diluted count; future buybacks and dilution are not separately modeled.", LBL),
        ("• Guidance anchors used: AMD FAD Nov-2025 (>35% rev CAGR, >60% DC CAGR, >$20 EPS in 3–5 yrs); NVIDIA "
         "≥$1T Blackwell+Rubin bookings through CY2027 and $3–4T AI-infra TAM by 2030; Snowflake Jun-2026 Investor "
         "Day ($10B product revenue FY29 reaffirmed, GAAP-profitable by Q4 FY28, SBC 41%→27% glidepath); Nutanix "
         "mid/high-teens growth by FY29 framework; Everpure (Pure Storage) Meta ramp + 2nd hyperscaler win (Aug-2026) "
         "ahead of its Sep-23-2026 analyst day; Sandisk Aug-2026 'In Focus' investor day (FY28–30: mid/high-teens "
         "growth, ~80% GM, ~75% op margin, ~50% FCF margin — haircut here for cycle risk); Kioxia capex-discipline "
         "plan (~¥470B/yr FY26–28); SK hynix HBM commitments and Samsung memory/foundry recovery (see tabs).", LBL),
        ("", LBL),
        ("CAVEATS", LBL_B),
        ("• Terminal value is typically 60–90% of EV in these models: value/share swings a lot with WACC, terminal "
         "growth and terminal margin — use the sensitivity grids and yellow levers, not the point estimates.", LBL),
        ("• For hyper-growth names (PLTR, NVDA, AMD, SNOW) a 5-year window structurally understates what the market "
         "is pricing: much of today's price is cash flows beyond year 5. Each tab's notes quantify the implied "
         "multiple gap. A negative 'upside' therefore reads as 'the market is paying for more than 5 years of this "
         "path', not necessarily 'sell'.", LBL),
        ("• Memory names are modeled with an explicit down-cycle in the 5-year path; timing of memory cycles is the "
         "single biggest swing factor and is essentially unforecastable to the year.", LBL),
        ("• KRW/JPY models are in local currency (₩bn / ¥bn; per-share in ₩/¥). Kioxia is pre-split (3-for-1 "
         "effective Oct 1, 2026 — divide by 3 after the split).", LBL),
        ("• Prices are the latest quotes retrievable on Aug 24, 2026 (a few are Aug 19–22 closes; noted per tab).", LBL),
        ("• This is an analytical exercise, not investment advice.", LBL),
    ]
    for i, (text, font) in enumerate(lines, start=1):
        style_cell(ws, f"A{i}", text, font, align=Alignment(wrap_text=True, vertical="top"))
        if font is LBL and len(text) > 100:
            ws.row_dimensions[i].height = 12 * (1 + len(text) // 105)
    return ws


def build(companies, path):
    wb = Workbook()
    wb.remove(wb.active)
    for co in companies:
        build_company_sheet(wb, co)
    build_summary(wb, companies)
    build_readme(wb)
    wb.active = 0
    wb.save(path)
    print(f"wrote {path}: {[s.title for s in wb.worksheets]}")
    for co in companies:
        s = python_dcf(co)
        print(f'  {co["ticker_disp"]:>8}: value/sh {s["vps"]:,.0f} vs price {co["price"]:,.0f} '
              f'({s["vps"] / co["price"] - 1:+.0%})  EV {s["ev"]:,.0f} {co["unit_short"]}')


if __name__ == "__main__":
    from company_data import COMPANIES
    build(COMPANIES, "dcf_models.xlsx")
