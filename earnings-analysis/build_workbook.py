#!/usr/bin/env python3
"""Build the two-sector (Software / Semiconductors) earnings comp workbook.

Rebuilds the user's Bloomberg BQL reference sheet ("QOQYOY.xlsx") into a full
earnings analysis book: quarterly revenue / adj. operating income / FCF pulled
via _xll.BQL spills (aligned by fiscal period offset), NTM estimate revisions
via _xll.BDP BEst fields, plus computed metric grids, a peer-comp dashboard and
a single-company "Focus" panel per sector.

Run: python3 build_workbook.py  ->  Earnings_Sector_Comp.xlsx (pre-postprocess)
"""
import re
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

OUT = "Earnings_Sector_Comp.xlsx"

# ---------------------------------------------------------------- ticker lists
SOFTWARE = """GOOG US Equity|MSFT US Equity|ORCL US Equity|AMD US Equity|SAP GY Equity|ARM US Equity|CRM US Equity|SHOP US Equity|PANW US Equity|CRWD US Equity|INTU US Equity|NOW US Equity|ADBE US Equity|CDNS US Equity|SNPS US Equity|NET US Equity|FTNT US Equity|CSU CN Equity|ADSK US Equity|SNOW US Equity|RBLX US Equity|DDOG US Equity|WDAY US Equity|AXON US Equity|ZM US Equity|VEEV US Equity|DSY FP Equity|P US Equity|ZS US Equity|TWLO US Equity|NTAP US Equity|MDB US Equity|PTC US Equity|SSNC US Equity|AKAM US Equity|TEAM US Equity|IOT US Equity|TRMB US Equity|CHKP US Equity|TYL US Equity|OKTA US Equity|BSY US Equity|HUBS US Equity|GDDY US Equity|GWRE US Equity|DT US Equity|FIG US Equity|NTNX US Equity|TTD US Equity|U US Equity|DOCN US Equity|DOCU US Equity|PCOR US Equity|PEGA US Equity|APPF US Equity|PAYC US Equity|FROG US Equity|PATH US Equity|PCTY US Equity|FSLY US Equity|S US Equity|DUOL US Equity|KXS CN Equity|BILL US Equity|CVLT US Equity|GTLB US Equity|MNDY US Equity|WK US Equity|BOX US Equity|RNG US Equity|TDC US Equity|BRZE US Equity|NCNO US Equity|SPSC US Equity|BL US Equity|TENB US Equity|VERX US Equity|BLKB US Equity|APPN US Equity|PLTR US Equity|RBRK US Equity|KVYO US Equity|AI US Equity|QLYS US Equity|VRNS US Equity|RDWR US Equity|ESTC US Equity|NICE US Equity|WIX US Equity|FRSH US Equity|AMPL US Equity|ASAN US Equity|CXM US Equity|RPD US Equity|PD US Equity|FIVN US Equity|EGHT US Equity|DOMO US Equity|SPT US Equity""".split("|")

SEMIS = """STM US Equity|ARM US Equity|1347 HK Equity|981 HK Equity|TSM US Equity|UMC US Equity|000660 KS Equity|SIMO US Equity|005930 KS Equity|285A JP Equity|PSTG US Equity|005935 KS Equity|1415 HK Equity|2382 HK Equity|ASX US Equity|042700 KS Equity|522 HK Equity|ASML NA Equity|BESI NA Equity|1478 HK Equity|1810 HK Equity|20 HK Equity|2018 HK Equity|2241 CH Equity|2371 CH Equity|2475 CH Equity|2533 HK Equity|2676 HK Equity|285 HK Equity|6613 HK Equity|9660 HK Equity|992 HK Equity|GDS US Equity|HIMX US Equity|SMHN GY Equity|SWKS US Equity|4919 TT Equity|SMCI US Equity|6789 TT Equity|5269 TT Equity|3035 TT Equity|NTAP US Equity|6770 TT Equity|2449 TT Equity|ACLS US Equity|3413 TT Equity|6533 TT Equity|4968 TT Equity|ALGM US Equity|WDC US Equity|3034 TT Equity|GFS US Equity|2379 TT Equity|OLED US Equity|ENTG US Equity|ON US Equity|DELL US Equity|QRVO US Equity|3592 TT Equity|4961 TT Equity|2344 TT Equity|RMBS US Equity|MCHP US Equity|ACMR US Equity|2303 TT Equity|AMKR US Equity|QCOM US Equity|2382 TT Equity|2458 TT Equity|2351 TT Equity|6239 TT Equity|COHR US Equity|5222 TT Equity|STX US Equity|8081 TT Equity|MPWR US Equity|3231 TT Equity|TER US Equity|3661 TT Equity|INTC US Equity|MU US Equity|6719 TT Equity|MRVL US Equity|2408 TT Equity|6415 TT Equity|6271 TT Equity|NXPI US Equity|6285 TT Equity|8150 TT Equity|SNPS US Equity|KLIC US Equity|KLAC US Equity|3006 TT Equity|CDNS US Equity|LRCX US Equity|6531 TT Equity|ADI US Equity|AMD US Equity|2313 TT Equity|3189 TT Equity|AMAT US Equity|3711 TT Equity|TXN US Equity|6799 TT Equity|NVDA US Equity|1560 TT Equity|AVGO US Equity|MSFT US Equity|AMZN US Equity|3443 TT Equity|6515 TT Equity|2360 TT Equity|GOOG US Equity|2317 TT Equity|3105 TT Equity|6147 TT Equity|2454 TT Equity|3491 TT Equity|8086 TT Equity|6488 TT Equity|8299 TT Equity|688052 CH Equity|300394 CH Equity|601138 CH Equity|300433 CH Equity|2330 TT Equity|688256 CH Equity|002371 CH Equity|688012 CH Equity|300782 CH Equity|300308 CH Equity|300661 CH Equity|688041 CH Equity|688981 CH Equity|603501 CH Equity|5347 TT Equity|688072 CH Equity|688347 CH Equity|3680 TT Equity""".split("|")

assert len(SOFTWARE) == len(set(SOFTWARE)), "dupes in software list"
assert len(SEMIS) == len(set(SEMIS)), "dupes in semis list"

# ---------------------------------------------------------------- style consts
F = "Arial"
BLUE = "0000FF"      # hardcoded inputs
GREEN = "008000"     # links to another sheet
BLACK = "000000"
WHITE = "FFFFFF"
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
BLOCK_FILL = PatternFill("solid", fgColor="D9E2F3")
EST_FILL = PatternFill("solid", fgColor="FFF2CC")
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
STAT_FILL = PatternFill("solid", fgColor="F2F2F2")
NOTE_GREY = "808080"

FMT_MM = '#,##0;(#,##0);"-"'
FMT_PCT = '0.0%;(0.0%);"-"'
FMT_PCT_SIGN = '+0.0%;(0.0%);"-"'
FMT_BPS = '+#,##0" bps";(#,##0)" bps";"-"'
FMT_X = '0.00"x"'
FMT_EPS = '0.00;(0.00);"-"'
FMT_DATE = 'dd-mmm-yy'
FMT_INT = '0'

def font(bold=False, color=BLACK, size=10, italic=False):
    return Font(name=F, bold=bold, color=color, size=size, italic=italic)

def hdr(ws, coord, text, comment=None, fill=HDR_FILL, color=WHITE, wrap=True):
    c = ws[coord]
    c.value = text
    c.font = font(bold=True, color=color)
    c.fill = fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    if comment:
        cm = Comment(comment, "Model notes")
        cm.width, cm.height = 320, 160
        c.comment = cm
    return c

def title(ws, text, sub=None):
    ws["A1"] = text
    ws["A1"].font = font(bold=True, size=13)
    if sub:
        ws["A2"] = sub
        ws["A2"].font = font(color=NOTE_GREY, size=9, italic=True)

def q(sheet):
    return f"'{sheet}'" if " " in sheet else sheet

# FPO ranges
DATA_FPO_LO, DATA_FPO_HI = -13, 2      # data sheets C..R
GRID_FPO_LO, GRID_FPO_HI = -9, 2       # metric grids C..N
DATA_START_ROW = 5                     # first ticker row on data sheets
def data_col(p):                       # column letter on data sheets for FPO p
    return get_column_letter(3 + (p - DATA_FPO_LO))
def grid_col(p):                       # column letter on metric grids for FPO p
    return get_column_letter(3 + (p - GRID_FPO_LO))
def fpo_label(p):
    if p > 0:
        return f"FPO +{p}E"
    return f"FPO {p}" if p else "FPO 0 (latest rpt)"

# ---------------------------------------------------------------- BQL builders
def bql_values(ticker_cell, field_expr):
    """One spilling BQL per ticker row: quarterly values FPO -13..+2, USD mm."""
    return (f'=_xll.BQL({ticker_cell}, "#itm().value as #v", '
            f'"#itm={field_expr}/1m","showids=f","transpose=t","showheaders=f")')

def bql_period(ticker_cell):
    """Latest reported quarter label, e.g. 2025 Q2."""
    return (f'=_xll.BQL({ticker_cell}, "#itm().period as #p", '
            f'"#itm=sales_rev_turn(fpt=q,fpo=0,ae=a)","showids=f","transpose=t","showheaders=f")')

REV_EXPR  = "sales_rev_turn(fpt=q,fpo=range(-13,2),ae=ae,fa_adjusted=y,currency='USD')"
OPINC_EXPR = "is_oper_inc(fpt=q,fpo=range(-13,2),ae=ae,fa_adjusted=y,currency='USD')"
FCF_EXPR  = "cf_free_cash_flow(fpt=q,fpo=range(-13,2),ae=ae,currency='USD')"

# ---------------------------------------------------------------- workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)

SECTORS = [
    ("SW", "Software", SOFTWARE, "ZM US Equity", "0070C0"),
    ("Semi", "Semiconductors", SEMIS, "NVDA US Equity", "ED7D31"),
]

# ============================================================ Tickers sheet
tk = wb.create_sheet("Tickers")
title(tk, "Ticker Lists (edit here — every other sheet reads these cells)",
      "Blue cells are inputs. Row order here = row order on every data/metric/dashboard sheet. "
      "If you add names, insert rows INSIDE the list on this sheet and copy formulas down on the sector sheets.")
TK_START = 5
for j, (code, name, lst, _, tab) in enumerate(SECTORS):
    col = 1 + j * 3   # A and D
    L = get_column_letter(col)
    hdr(tk, f"{L}4", f"{name} ({len(lst)} tickers)")
    tk.row_dimensions[4].height = 20
    for i, t in enumerate(lst):
        c = tk.cell(row=TK_START + i, column=col, value=t)
        c.font = font(color=BLUE)
    n = tk.cell(row=TK_START + len(lst) + 1, column=col,
                value=f'=COUNTA({L}{TK_START}:{L}{TK_START + len(lst) - 1})&" tickers"')
    n.font = font(italic=True, color=NOTE_GREY)
    tk.column_dimensions[L].width = 20
tk.sheet_properties.tabColor = "A6A6A6"
tk.freeze_panes = "A5"
tick_addr = {}  # sector code -> (col letter, list)
tick_addr["SW"] = ("A", SOFTWARE)
tick_addr["Semi"] = ("D", SEMIS)

# ============================================================ data sheets
DATA_SHEETS = [
    ("Rev", "Quarterly Adjusted Revenue (USD mm)", REV_EXPR,
     "BQL: sales_rev_turn, fpt=q, ae=ae (actual through FPO 0, consensus estimate for FPO +1/+2), "
     "fa_adjusted=y (Bloomberg-adjusted), currency='USD', scaled /1m."),
    ("OpInc", "Quarterly Adjusted Operating Income (USD mm)", OPINC_EXPR,
     "BQL: is_oper_inc, fpt=q, ae=ae, fa_adjusted=y, currency='USD', /1m."),
    ("FCF", "Quarterly Free Cash Flow (USD mm)", FCF_EXPR,
     "BQL: cf_free_cash_flow, fpt=q, ae=ae, currency='USD', /1m. "
     "FCF estimates are sparse for smaller names — blanks/#N/A in FPO +1/+2 are normal."),
]

for code, sector, lst, focus_default, tab in SECTORS:
    tcol, _ = tick_addr[code]
    n = len(lst)
    end_row = DATA_START_ROW + n - 1
    for suffix, label, expr, srcnote in DATA_SHEETS:
        ws = wb.create_sheet(f"{code} {suffix}")
        ws.sheet_properties.tabColor = tab
        title(ws, f"{sector} — {label}",
              "Columns are aligned by fiscal period offset (FPO): FPO 0 = each company's most recently REPORTED "
              "quarter, so fiscal calendars differ across rows (e.g. ZM's FY ends Jan). FPO +1/+2 are consensus estimates (shaded).")
        ws["A3"] = f"Source: Bloomberg — {srcnote} Refresh on a Bloomberg terminal (Bloomberg ribbon > Refresh All)."
        ws["A3"].font = font(color=NOTE_GREY, size=9, italic=True)
        hdr(ws, "A4", "Ticker")
        hdr(ws, "B4", "Latest Rpt Qtr",
            comment="Pulled by BQL: period label of FPO 0 (the last reported fiscal quarter) for THIS company. "
                    "Use it to see which calendar quarter each row's FPO 0 refers to.")
        for p in range(DATA_FPO_LO, DATA_FPO_HI + 1):
            cl = data_col(p)
            hdr(ws, f"{cl}4", fpo_label(p))
            if p > 0:
                for r in range(DATA_START_ROW, end_row + 1):
                    ws[f"{cl}{r}"].fill = EST_FILL
        ws["A4"].comment = Comment(
            "Tickers link to the Tickers sheet (single source of truth). Edit lists there.", "Model notes", width=280, height=90)
        for i in range(n):
            r = DATA_START_ROW + i
            a = ws[f"A{r}"]
            a.value = f"=Tickers!${tcol}${TK_START + i}"
            a.font = font(color=GREEN)
            if suffix == "Rev":
                b = ws[f"B{r}"]
                b.value = ArrayFormula(f"B{r}", bql_period(f"$A{r}"))
                b.font = font()
            else:
                b = ws[f"B{r}"]
                b.value = f"={q(code + ' Rev')}!$B{r}"
                b.font = font(color=GREEN)
            ws[f"C{r}"] = ArrayFormula(f"C{r}", bql_values(f"$A{r}", expr))
            for p in range(DATA_FPO_LO, DATA_FPO_HI + 1):
                ws[f"{data_col(p)}{r}"].number_format = FMT_MM
                ws[f"{data_col(p)}{r}"].font = font()
        ws.column_dimensions["A"].width = 17
        ws.column_dimensions["B"].width = 12
        for p in range(DATA_FPO_LO, DATA_FPO_HI + 1):
            ws.column_dimensions[data_col(p)].width = 9.5
        ws.row_dimensions[4].height = 28
        ws.freeze_panes = "C5"

# ============================================================ revisions sheets
REVISION_ROW0 = 6
for code, sector, lst, focus_default, tab in SECTORS:
    tcol, _ = tick_addr[code]
    n = len(lst)
    ws = wb.create_sheet(f"{code} Revisions")
    ws.sheet_properties.tabColor = tab
    title(ws, f"{sector} — NTM Consensus Estimate Revisions",
          "NTM (blended-forward 12M, BEst 1BF) consensus revenue and EPS: current vs 1M / 3M / 6M ago. "
          "Revisions are % changes, so mixed reporting currencies do not distort them.")
    ws["A3"] = ("Source: Bloomberg BDP — BEST_SALES / BEST_EPS with BEST_FPERIOD_OVERRIDE=1BF; "
                "historical as-of via BEST_DATA_DATE_OVERRIDE (dates in row 4, driven by TODAY()).")
    ws["A3"].font = font(color=NOTE_GREY, size=9, italic=True)
    labels = ["Current", "1M ago", "3M ago", "6M ago"]
    date_cells = {}
    for k, lab in enumerate(labels):
        cl = get_column_letter(3 + k)  # C..F
        ws[f"{cl}4"] = "=TODAY()" if k == 0 else f"=EDATE($C$4,-{[0, 1, 3, 6][k]})"
        ws[f"{cl}4"].number_format = FMT_DATE
        ws[f"{cl}4"].font = font(bold=True)
        date_cells[lab] = f"${cl}$4"
    ws["B4"] = "As-of dates:"
    ws["B4"].font = font(bold=True)
    hdr(ws, "A5", "Ticker")
    hdr(ws, "B5", "")
    hdr(ws, "J5", "")
    ws.row_dimensions[5].height = 30
    heads = [("C", "NTM Rev\n(loc mm)"), ("D", "NTM Rev\n1M ago"), ("E", "NTM Rev\n3M ago"), ("F", "NTM Rev\n6M ago"),
             ("G", "Rev rev\nΔ1M %"), ("H", "Rev rev\nΔ3M %"), ("I", "Rev rev\nΔ6M %"),
             ("K", "NTM EPS"), ("L", "NTM EPS\n1M ago"), ("M", "NTM EPS\n3M ago"), ("N", "NTM EPS\n6M ago"),
             ("O", "EPS rev\nΔ1M %"), ("P", "EPS rev\nΔ3M %"), ("Q", "EPS rev\nΔ6M %")]
    for cl, t in heads:
        hdr(ws, f"{cl}5", t)
    ws["G5"].comment = Comment("Δ = current NTM consensus / consensus as-of the earlier date − 1. "
                               "Positive = upward revisions.", "Model notes", width=280, height=90)
    ws["O5"].comment = Comment("EPS revision % is unreliable when EPS is near zero or negative "
                               "(sign flips) — read together with the revenue revision.", "Model notes", width=280, height=110)
    for i in range(n):
        r = REVISION_ROW0 + i
        a = ws[f"A{r}"]
        a.value = f"=Tickers!${tcol}${TK_START + i}"
        a.font = font(color=GREEN)
        for fld, base in (("BEST_SALES", "C"), ("BEST_EPS", "K")):
            cols = [get_column_letter(ord(base) - 64 + k) for k in range(4)]
            for k, cl in enumerate(cols):
                if k == 0:
                    ws[f"{cl}{r}"] = f'=_xll.BDP($A{r},"{fld}","BEST_FPERIOD_OVERRIDE","1BF")'
                else:
                    dc = date_cells[labels[k]]
                    # YEAR/MONTH/DAY + "0000"/"00" is locale-proof (TEXT(date,"YYYYMMDD")
                    # breaks on non-English Excel locales, e.g. German JJJJ)
                    dstr = (f'TEXT(YEAR({dc}),"0000")&TEXT(MONTH({dc}),"00")&TEXT(DAY({dc}),"00")')
                    ws[f"{cl}{r}"] = (f'=_xll.BDP($A{r},"{fld}","BEST_FPERIOD_OVERRIDE","1BF",'
                                      f'"BEST_DATA_DATE_OVERRIDE",{dstr})')
                ws[f"{cl}{r}"].number_format = FMT_MM if fld == "BEST_SALES" else FMT_EPS
                ws[f"{cl}{r}"].font = font()
            dcols = [get_column_letter(ord(base) - 64 + 4 + k) for k in range(3)]
            for k, cl in enumerate(dcols):
                ago = cols[k + 1]
                cur = cols[0]
                ws[f"{cl}{r}"] = (f'=IFERROR(IF(AND(ISNUMBER(${cur}{r}),ISNUMBER({ago}{r}),{ago}{r}<>0),'
                                  f'${cur}{r}/{ago}{r}-1,""),"")')
                ws[f"{cl}{r}"].number_format = FMT_PCT_SIGN
                ws[f"{cl}{r}"].font = font()
    ws.column_dimensions["A"].width = 17
    for cl in "CDEFGHIKLMNOPQ":
        ws.column_dimensions[cl].width = 10.5
    ws.column_dimensions["J"].width = 2
    ws.freeze_panes = "B6"
    last = REVISION_ROW0 + n - 1
    for rng in (f"G{REVISION_ROW0}:I{last}", f"O{REVISION_ROW0}:Q{last}"):
        ws.conditional_formatting.add(rng, ColorScaleRule(
            start_type="num", start_value=-0.10, start_color="F8696B",
            mid_type="num", mid_value=0, mid_color="FFFFFF",
            end_type="num", end_value=0.10, end_color="63BE7B"))

# ============================================================ metric grids
BLOCKS = [
    ("rev", "Revenue (USD mm)", FMT_MM,
     "Quarterly Bloomberg-adjusted revenue, USD mm — error-guarded mirror of the Rev data sheet "
     "(blank where Bloomberg returns #N/A, e.g. pre-IPO quarters)."),
    ("rev_yoy", "Revenue growth YoY %", FMT_PCT,
     "Rev(p) / Rev(p−4) − 1. Blank when the base quarter is missing or ≤ 0."),
    ("rev_qoq", "Revenue growth QoQ %", FMT_PCT,
     "Rev(p) / Rev(p−1) − 1."),
    ("add_yoy", "Revenue $ added YoY (USD mm)", FMT_MM,
     "Rev(p) − Rev(p−4): incremental annualized-quarter revenue vs the same quarter last year."),
    ("add_qoq", "Revenue $ added QoQ (USD mm)", FMT_MM,
     "Rev(p) − Rev(p−1): sequential $ added."),
    ("opm", "Adj. operating margin %", FMT_PCT,
     "OpInc(p) / Rev(p)."),
    ("opm_yoy", "Op margin expansion YoY (bps)", FMT_BPS,
     "[OpInc(p)/Rev(p) − OpInc(p−4)/Rev(p−4)] × 10,000."),
    ("oi_yoy", "Adj. op income growth YoY %", FMT_PCT,
     "OpInc(p) / OpInc(p−4) − 1. Blank when the base is ≤ 0 (growth % vs a negative base is meaningless)."),
    ("incr_opm", "Incremental op margin YoY % (operating leverage)", FMT_PCT,
     "[OpInc(p) − OpInc(p−4)] / [Rev(p) − Rev(p−4)]: margin on each new $ of revenue. "
     "Blank when |Δrevenue| < $1mm. Above the current margin = operating leverage."),
    ("fcf_ltm", "FCF margin, LTM %", FMT_PCT,
     "Σ FCF(p−3..p) / Σ Rev(p−3..p). LTM smooths quarterly FCF lumpiness (working-capital timing)."),
    ("fcf_ltm_yoy", "FCF margin expansion, LTM YoY (bps)", FMT_BPS,
     "LTM FCF margin(p) − LTM FCF margin(p−4). Needs 8 trailing quarters, so starts at FPO −6."),
]

block_rows = {}   # (code, key) -> (first_data_row, last_data_row, ticker_col_range)

for code, sector, lst, focus_default, tab in SECTORS:
    n = len(lst)
    ws = wb.create_sheet(f"{code} Metrics")
    ws.sheet_properties.tabColor = tab
    rev = q(f"{code} Rev")
    opi = q(f"{code} OpInc")
    fcf = q(f"{code} FCF")
    title(ws, f"{sector} — Metric Grids (by fiscal period offset)",
          "All inputs from the BQL data sheets. FPO 0 = latest reported quarter per company; +1/+2 = consensus estimates (shaded). "
          "Blank cells = insufficient history / no estimate. Hover column headers and block titles for definitions.")
    r = 4
    for key, name, fmt, desc in BLOCKS:
        tcell = ws.cell(row=r, column=1, value=name)
        tcell.font = font(bold=True, size=11)
        tcell.fill = BLOCK_FILL
        for cidx in range(2, 15):
            bc = ws.cell(row=r, column=cidx)
            bc.fill = BLOCK_FILL
            bc.font = font()
        tcell.comment = Comment(desc, "Model notes", width=340, height=130)
        hr = r + 1
        ws.row_dimensions[hr].height = 28
        hdr(ws, f"A{hr}", "Ticker")
        hdr(ws, f"B{hr}", "Latest Rpt Qtr")
        for p in range(GRID_FPO_LO, GRID_FPO_HI + 1):
            hdr(ws, f"{grid_col(p)}{hr}", fpo_label(p))
        r0 = hr + 1
        for i in range(n):
            rr = r0 + i
            dr = DATA_START_ROW + i
            ws[f"A{rr}"] = f"={rev}!$A{dr}"
            ws[f"A{rr}"].font = font(color=GREEN)
            ws[f"B{rr}"] = f"={rev}!$B{dr}"
            ws[f"B{rr}"].font = font(color=GREEN)
            for p in range(GRID_FPO_LO, GRID_FPO_HI + 1):
                cl = grid_col(p)
                cp, cp1, cp4 = data_col(p), data_col(p - 1), data_col(p - 4)
                R = lambda c: f"{rev}!{c}{dr}"
                O = lambda c: f"{opi}!{c}{dr}"
                C = lambda c: f"{fcf}!{c}{dr}"
                if key == "rev":
                    f_ = f'=IFERROR(IF(ISNUMBER({R(cp)}),{R(cp)},""),"")'
                elif key == "rev_yoy":
                    f_ = (f'=IFERROR(IF(AND(ISNUMBER({R(cp)}),ISNUMBER({R(cp4)}),{R(cp4)}>0),'
                          f'{R(cp)}/{R(cp4)}-1,""),"")')
                elif key == "rev_qoq":
                    f_ = (f'=IFERROR(IF(AND(ISNUMBER({R(cp)}),ISNUMBER({R(cp1)}),{R(cp1)}>0),'
                          f'{R(cp)}/{R(cp1)}-1,""),"")')
                elif key == "add_yoy":
                    f_ = (f'=IFERROR(IF(AND(ISNUMBER({R(cp)}),ISNUMBER({R(cp4)})),'
                          f'{R(cp)}-{R(cp4)},""),"")')
                elif key == "add_qoq":
                    f_ = (f'=IFERROR(IF(AND(ISNUMBER({R(cp)}),ISNUMBER({R(cp1)})),'
                          f'{R(cp)}-{R(cp1)},""),"")')
                elif key == "opm":
                    f_ = (f'=IFERROR(IF(AND(ISNUMBER({O(cp)}),ISNUMBER({R(cp)}),{R(cp)}>0),'
                          f'{O(cp)}/{R(cp)},""),"")')
                elif key == "opm_yoy":
                    f_ = (f'=IFERROR(IF(AND(ISNUMBER({O(cp)}),ISNUMBER({R(cp)}),ISNUMBER({O(cp4)}),'
                          f'ISNUMBER({R(cp4)}),{R(cp)}>0,{R(cp4)}>0),'
                          f'({O(cp)}/{R(cp)}-{O(cp4)}/{R(cp4)})*10000,""),"")')
                elif key == "oi_yoy":
                    f_ = (f'=IFERROR(IF(AND(ISNUMBER({O(cp)}),ISNUMBER({O(cp4)}),{O(cp4)}>0),'
                          f'{O(cp)}/{O(cp4)}-1,""),"")')
                elif key == "incr_opm":
                    f_ = (f'=IFERROR(IF(AND(ISNUMBER({O(cp)}),ISNUMBER({O(cp4)}),ISNUMBER({R(cp)}),'
                          f'ISNUMBER({R(cp4)}),ABS({R(cp)}-{R(cp4)})>=1),'
                          f'({O(cp)}-{O(cp4)})/({R(cp)}-{R(cp4)}),""),"")')
                elif key == "fcf_ltm":
                    cpm3 = data_col(p - 3)
                    f_ = (f'=IFERROR(IF(SUM({rev}!{cpm3}{dr}:{cp}{dr})<=0,"",'
                          f'SUM({fcf}!{cpm3}{dr}:{cp}{dr})/SUM({rev}!{cpm3}{dr}:{cp}{dr})),"")')
                elif key == "fcf_ltm_yoy":
                    if p - 7 < DATA_FPO_LO:
                        f_ = '=""'   # needs 8 trailing qtrs; keep as formula so INDEX on Focus returns blank, not 0
                    else:
                        fr0, fr1, _ = block_rows[(code, "fcf_ltm")]
                        ltm_now = f"${grid_col(p)}${fr0 + i}"
                        ltm_prior = f"${grid_col(p - 4)}${fr0 + i}"
                        # p-4 may be off-grid (< -9): compute prior-year LTM from data directly
                        if p - 4 < GRID_FPO_LO:
                            c_hi, c_lo = data_col(p - 4), data_col(p - 7)
                            ltm_prior_expr = (f'IF(SUM({rev}!{c_lo}{dr}:{c_hi}{dr})<=0,"",'
                                              f'SUM({fcf}!{c_lo}{dr}:{c_hi}{dr})/SUM({rev}!{c_lo}{dr}:{c_hi}{dr}))')
                        else:
                            ltm_prior_expr = ltm_prior
                        f_ = (f'=IFERROR(IF(OR({ltm_now}="",{ltm_prior_expr}=""),"",'
                              f'({ltm_now}-{ltm_prior_expr})*10000),"")')
                cc = ws[f"{cl}{rr}"]
                cc.value = f_
                cc.number_format = fmt
                cc.font = font()
                if p > 0:
                    cc.fill = EST_FILL
        last = r0 + n - 1
        block_rows[(code, key)] = (r0, last, None)
        ws.conditional_formatting.add(
            f"C{r0}:N{last}",
            ColorScaleRule(start_type="percentile", start_value=10, start_color="F8696B",
                           mid_type="percentile", mid_value=50, mid_color="FFEB84",
                           end_type="percentile", end_value=90, end_color="63BE7B"))
        r = last + 2
    ws.column_dimensions["A"].width = 17
    ws.column_dimensions["B"].width = 12
    for p in range(GRID_FPO_LO, GRID_FPO_HI + 1):
        ws.column_dimensions[grid_col(p)].width = 10
    ws.freeze_panes = "C2"

# ============================================================ dashboards
DASH_ROW_HDR = 4
DASH_STATS = [("Peer Median", "MEDIAN({rng})"),
              ("Top Quartile (75th)", "QUARTILE({rng},3)"),
              ("Bottom Quartile (25th)", "QUARTILE({rng},1)")]
DASH_ROW0 = DASH_ROW_HDR + len(DASH_STATS) + 1   # 8

for code, sector, lst, focus_default, tab in SECTORS:
    n = len(lst)
    ws = wb.create_sheet(f"{code} Dashboard")
    ws.sheet_properties.tabColor = tab
    rev = q(f"{code} Rev")
    met = q(f"{code} Metrics")
    rvs = q(f"{code} Revisions")
    title(ws, f"{sector} — Earnings Comp Dashboard (latest reported quarter)",
          "USD mm unless noted. Every row reads the same-numbered row on the data/metric sheets. "
          "FPO 0 = each company's own latest reported quarter. Estimate-based columns note (Est).")
    cols = [
        ("A", "Ticker", None, None),
        ("B", "Latest Rpt Qtr", None, None),
        ("C", "Revenue LQ (USD mm)", FMT_MM, "Latest reported quarter, Bloomberg-adjusted, USD."),
        ("D", "Rev YoY %", FMT_PCT, None),
        ("E", "Rev YoY % (prior qtr)", FMT_PCT, "Same metric one quarter earlier (FPO −1) — the base for acceleration."),
        ("F", "Accel / (Decel) (pp)", FMT_PCT_SIGN, "YoY growth this quarter minus YoY growth last quarter, in percentage points."),
        ("G", "Rev QoQ %", FMT_PCT, None),
        ("H", "$ Added QoQ (mm)", FMT_MM, None),
        ("I", "$ Added YoY (mm)", FMT_MM, None),
        ("J", "Next Qtr YoY % (Est)", FMT_PCT, "Consensus FPO +1 revenue vs the same quarter last year (FPO −3)."),
        ("K", "Op Margin %", FMT_PCT, "Adjusted operating margin, latest reported quarter."),
        ("L", "OpM Δ YoY (bps)", FMT_BPS, None),
        ("M", "Op Inc YoY %", FMT_PCT, None),
        ("N", "Incr OpM % (Op Leverage)", FMT_PCT, "Incremental operating margin: ΔOpInc / ΔRev YoY. Above current margin = leverage."),
        ("O", "Op Lev (x)", FMT_X, "Op income YoY growth ÷ revenue YoY growth. Shown only when both are positive."),
        ("P", "FCF Margin LTM %", FMT_PCT, None),
        ("Q", "FCF Margin Δ YoY (bps)", FMT_BPS, "LTM FCF margin now vs LTM FCF margin a year ago."),
        ("R", "NTM Rev Rev Δ1M %", FMT_PCT_SIGN, "Change in NTM (1BF) consensus revenue over the last month."),
        ("S", "NTM Rev Rev Δ3M %", FMT_PCT_SIGN, None),
        ("T", "NTM Rev Rev Δ6M %", FMT_PCT_SIGN, None),
        ("U", "NTM EPS Rev Δ1M %", FMT_PCT_SIGN, None),
        ("V", "NTM EPS Rev Δ3M %", FMT_PCT_SIGN, None),
        ("W", "NTM EPS Rev Δ6M %", FMT_PCT_SIGN, None),
        ("X", "Rank: Rev YoY", FMT_INT, "1 = fastest YoY growth in the peer set."),
        ("Y", "Rank: NTM Rev Δ3M", FMT_INT, "1 = biggest upward 3-month NTM revenue revision."),
    ]
    for cl, t, _, cmt in cols:
        hdr(ws, f"{cl}{DASH_ROW_HDR}", t, comment=cmt)
    last = DASH_ROW0 + n - 1
    for k, (lab, fx) in enumerate(DASH_STATS):
        sr = DASH_ROW_HDR + 1 + k
        ws[f"A{sr}"] = lab
        ws[f"A{sr}"].font = font(bold=True, italic=True)
        ws[f"A{sr}"].fill = STAT_FILL
        ws[f"B{sr}"].fill = STAT_FILL
        for cl, t, fmt, _ in cols[2:]:
            c = ws[f"{cl}{sr}"]
            c.font = font(italic=True)
            c.fill = STAT_FILL
            if cl in ("X", "Y"):
                continue          # no stats over rank columns, but keep the band unbroken
            rng = f"{cl}{DASH_ROW0}:{cl}{last}"
            c.value = f'=IFERROR({fx.format(rng=rng)},"")'
            c.number_format = fmt
    for i in range(n):
        r = DASH_ROW0 + i
        dr = DATA_START_ROW + i
        def mref(key, p):
            r0, _, _ = block_rows[(code, key)]
            return f"{met}!{grid_col(p)}{r0 + i}"
        vr = REVISION_ROW0 + i
        ws[f"A{r}"] = f"={rev}!$A{dr}"
        ws[f"A{r}"].font = font(color=GREEN)
        ws[f"B{r}"] = f"={rev}!$B{dr}"
        ws[f"B{r}"].font = font(color=GREEN)
        vals = {
            "C": f'=IF(ISNUMBER({rev}!{data_col(0)}{dr}),{rev}!{data_col(0)}{dr},"")',
            "D": f"={mref('rev_yoy', 0)}",
            "E": f"={mref('rev_yoy', -1)}",
            "F": f'=IF(OR(D{r}="",E{r}=""),"",D{r}-E{r})',
            "G": f"={mref('rev_qoq', 0)}",
            "H": f"={mref('add_qoq', 0)}",
            "I": f"={mref('add_yoy', 0)}",
            "J": f"={mref('rev_yoy', 1)}",
            "K": f"={mref('opm', 0)}",
            "L": f"={mref('opm_yoy', 0)}",
            "M": f"={mref('oi_yoy', 0)}",
            "N": f"={mref('incr_opm', 0)}",
            "O": f'=IFERROR(IF(AND(N(D{r})>0,N(M{r})>0),M{r}/D{r},""),"")',
            "P": f"={mref('fcf_ltm', 0)}",
            "Q": f"={mref('fcf_ltm_yoy', 0)}",
            "R": f"={rvs}!G{vr}",
            "S": f"={rvs}!H{vr}",
            "T": f"={rvs}!I{vr}",
            "U": f"={rvs}!O{vr}",
            "V": f"={rvs}!P{vr}",
            "W": f"={rvs}!Q{vr}",
            "X": f'=IFERROR(RANK(D{r},D${DASH_ROW0}:D${last}),"")',
            "Y": f'=IFERROR(RANK(S{r},S${DASH_ROW0}:S${last}),"")',
        }
        for cl, t, fmt, _ in cols[2:]:
            c = ws[f"{cl}{r}"]
            c.value = vals[cl]
            if fmt:
                c.number_format = fmt
            c.font = font(color=BLACK if cl in ("F", "O", "X", "Y") else GREEN)
    for cl, w in [("A", 17), ("B", 11)] + [(c, 11) for c, _, _, _ in cols[2:]]:
        ws.column_dimensions[cl].width = w
    ws.freeze_panes = f"C{DASH_ROW0}"
    ws.row_dimensions[DASH_ROW_HDR].height = 42
    for cl in ["D", "F", "J", "L", "N", "Q", "R", "S", "T", "U", "V", "W"]:
        ws.conditional_formatting.add(
            f"{cl}{DASH_ROW0}:{cl}{last}",
            ColorScaleRule(start_type="percentile", start_value=10, start_color="F8696B",
                           mid_type="percentile", mid_value=50, mid_color="FFEB84",
                           end_type="percentile", end_value=90, end_color="63BE7B"))

# ============================================================ focus sheets
FOCUS_METRICS = [
    ("Revenue (USD mm)", "rev", FMT_MM, "grid"),
    ("Revenue growth YoY %", "rev_yoy", FMT_PCT, "grid"),
    ("Revenue growth QoQ %", "rev_qoq", FMT_PCT, "grid"),
    ("$ Added YoY (USD mm)", "add_yoy", FMT_MM, "grid"),
    ("$ Added QoQ (USD mm)", "add_qoq", FMT_MM, "grid"),
    ("Op margin %", "opm", FMT_PCT, "grid"),
    ("Op margin Δ YoY (bps)", "opm_yoy", FMT_BPS, "grid"),
    ("Op income YoY %", "oi_yoy", FMT_PCT, "grid"),
    ("Incremental OpM % (leverage)", "incr_opm", FMT_PCT, "grid"),
    ("FCF margin LTM %", "fcf_ltm", FMT_PCT, "grid"),
    ("FCF margin Δ YoY (bps)", "fcf_ltm_yoy", FMT_BPS, "grid"),
]

for code, sector, lst, focus_default, tab in SECTORS:
    n = len(lst)
    ws = wb.create_sheet(f"{code} Focus")
    ws.sheet_properties.tabColor = tab
    rev = q(f"{code} Rev")
    met = q(f"{code} Metrics")
    rvs = q(f"{code} Revisions")
    title(ws, f"{sector} — Single-Name vs Peers",
          "Type any ticker from the Tickers sheet into the yellow cell. Each metric shows the company, "
          "the peer-set median (computed over the full list, selected name included), and the gap. "
          "Columns aligned by FPO (company's own fiscal quarters).")
    ws["A4"] = "Company (edit):"
    ws["A4"].font = font(bold=True)
    ic = ws["B4"]
    ic.value = focus_default
    ic.font = font(bold=True, color=BLUE)
    ic.fill = YELLOW_FILL
    dlast = DATA_START_ROW + n - 1
    ws["C4"] = (f'=IFERROR("Latest reported qtr: "&INDEX({rev}!$B${DATA_START_ROW}:$B${dlast},'
                f'MATCH($B$4,{rev}!$A${DATA_START_ROW}:$A${dlast},0)),"⚠ ticker not found in Tickers list")')
    ws["C4"].font = font(italic=True, color=NOTE_GREY)
    hr = 6
    hdr(ws, f"A{hr}", "Metric")
    hdr(ws, f"B{hr}", "Series")
    for p in range(GRID_FPO_LO, GRID_FPO_HI + 1):
        hdr(ws, f"{grid_col(p)}{hr}", fpo_label(p))
    ws.row_dimensions[hr].height = 28
    r = hr + 1
    for name, key, fmt, kind in FOCUS_METRICS:
        for series in ("Company", "Peer median", "vs median"):
            ws[f"A{r}"] = name if series == "Company" else ""
            ws[f"A{r}"].font = font(bold=(series == "Company"))
            ws[f"B{r}"] = series
            ws[f"B{r}"].font = font(italic=(series != "Company"), size=9 if series != "Company" else 10)
            for p in range(GRID_FPO_LO, GRID_FPO_HI + 1):
                cl = grid_col(p)
                if kind == "data":
                    src_col = data_col(p)
                    col_rng = f"{rev}!${src_col}${DATA_START_ROW}:${src_col}${dlast}"
                    match_rng = f"{rev}!$A${DATA_START_ROW}:$A${dlast}"
                else:
                    r0, r1, _ = block_rows[(code, key)]
                    col_rng = f"{met}!${cl}${r0}:${cl}${r1}"
                    match_rng = f"{met}!$A${r0}:$A${r1}"
                if series == "Company":
                    f_ = f'=IFERROR(INDEX({col_rng},MATCH($B$4,{match_rng},0)),"")'
                elif series == "Peer median":
                    f_ = f'=IFERROR(MEDIAN({col_rng}),"")'
                else:
                    f_ = f'=IF(OR({cl}{r-2}="",{cl}{r-1}=""),"",{cl}{r-2}-{cl}{r-1})'
                c = ws[f"{cl}{r}"]
                c.value = f_
                c.number_format = fmt
                c.font = font(size=9 if series != "Company" else 10,
                              italic=(series == "vs median"))
                if p > 0:
                    c.fill = EST_FILL
            r += 1
        r += 1
    # NTM revisions mini-block
    ws[f"A{r}"] = "NTM estimate revisions"
    ws[f"A{r}"].font = font(bold=True, size=11)
    ws[f"A{r}"].fill = BLOCK_FILL
    for cidx in range(2, 6):
        bc = ws.cell(row=r, column=cidx)
        bc.fill = BLOCK_FILL
        bc.font = font()
    r += 1
    hdr(ws, f"A{r}", "Metric")
    hdr(ws, f"B{r}", "Series")
    hdr(ws, f"C{r}", "Δ1M %")
    hdr(ws, f"D{r}", "Δ3M %")
    hdr(ws, f"E{r}", "Δ6M %")
    r += 1
    rlast = REVISION_ROW0 + n - 1
    for name, colset in [("NTM Revenue revision", ("G", "H", "I")),
                         ("NTM EPS revision", ("O", "P", "Q"))]:
        for series in ("Company", "Peer median", "vs median"):
            ws[f"A{r}"] = name if series == "Company" else ""
            ws[f"A{r}"].font = font(bold=(series == "Company"))
            ws[f"B{r}"] = series
            ws[f"B{r}"].font = font(italic=(series != "Company"), size=9 if series != "Company" else 10)
            for k, cl in enumerate(("C", "D", "E")):
                src = colset[k]
                col_rng = f"{rvs}!${src}${REVISION_ROW0}:${src}${rlast}"
                match_rng = f"{rvs}!$A${REVISION_ROW0}:$A${rlast}"
                if series == "Company":
                    f_ = f'=IFERROR(INDEX({col_rng},MATCH($B$4,{match_rng},0)),"")'
                elif series == "Peer median":
                    f_ = f'=IFERROR(MEDIAN({col_rng}),"")'
                else:
                    f_ = f'=IF(OR({cl}{r-2}="",{cl}{r-1}=""),"",{cl}{r-2}-{cl}{r-1})'
                c = ws[f"{cl}{r}"]
                c.value = f_
                c.number_format = FMT_PCT_SIGN
                c.font = font(size=9 if series != "Company" else 10,
                              italic=(series == "vs median"))
            r += 1
        r += 1
    ws.column_dimensions["A"].width = 27
    ws.column_dimensions["B"].width = 12
    for p in range(GRID_FPO_LO, GRID_FPO_HI + 1):
        ws.column_dimensions[grid_col(p)].width = 10
    ws.freeze_panes = "C7"

# ============================================================ README
rd = wb.create_sheet("README", 0)
rd.sheet_properties.tabColor = "A6A6A6"
rows = [
    ("Earnings by Sector — Peer Comp Workbook", "T1"),
    (f"Software ({len(SOFTWARE)} names) and Semiconductors ({len(SEMIS)} names) quarterly earnings analysis vs peers: revenue growth, "
     "$ added QoQ/YoY, margin expansion, operating leverage, FCF margin expansion, and NTM estimate revisions.", "sub"),
    ("", None),
    ("HOW TO REFRESH", "H"),
    ("1. Open this file in Excel on a machine with the Bloomberg Add-in (terminal) logged in.", None),
    ("2. Bloomberg ribbon > Refresh All (or Ctrl+Alt+F9). The BQL spills and BDP calls will populate; the first "
     "refresh pulls ~950 BQL queries + ~1,900 BDP fields, so give it a few minutes.", None),
    ("3. All analysis sheets (Metrics, Dashboard, Focus) recalculate automatically from the data sheets.", None),
    ("", None),
    ("SHEET GUIDE", "H"),
    ("Tickers — the only input sheet: both peer lists. Row order here drives every other sheet.", None),
    ("SW/Semi Rev, OpInc, FCF — raw quarterly data via one BQL spill per ticker (FPO −13 … +2, USD mm).", None),
    ("SW/Semi Revisions — NTM (1BF) consensus revenue & EPS now vs 1M/3M/6M ago, and the % revisions.", None),
    ("SW/Semi Metrics — computed grids: YoY/QoQ growth, $ added, op margin & expansion, op income growth, "
     "incremental margin (operating leverage), LTM FCF margin & expansion.", None),
    ("SW/Semi Dashboard — one row per name, latest-quarter snapshot of every metric + revision columns, with "
     "peer median / quartiles on top and color scales.", None),
    ("SW/Semi Focus — type one ticker (e.g. ZM US Equity) and see its full time series vs the peer median.", None),
    ("", None),
    ("METHODOLOGY / CONVENTIONS", "H"),
    ("Alignment: columns are fiscal period offsets (FPO). FPO 0 = each company's most recently reported quarter, "
     "FPO +1/+2 = consensus estimates (ae=ae). Fiscal calendars differ (ZM FY ends Jan), so FPO alignment — not "
     "calendar quarters — is used for the peer comp, matching the original sheet's design.", None),
    ("Currency: all BQL pulls request currency='USD' so cross-listed names (TWD/KRW/JPY/EUR/HKD/CNY) are "
     "comparable in $ terms. If your BQL version rejects the currency parameter, delete \",currency='USD'\" from "
     "the formulas in column C of the data sheets (values then come in filing currency).", None),
    ("Fields: revenue = sales_rev_turn (fa_adjusted=y), op income = is_oper_inc (fa_adjusted=y), "
     "FCF = cf_free_cash_flow — all fpt=q, /1m. NTM = BDP BEST_SALES / BEST_EPS with BEST_FPERIOD_OVERRIDE=1BF; "
     "historical consensus via BEST_DATA_DATE_OVERRIDE at TODAY()−1M/3M/6M (dates recompute daily).", None),
    ("Operating leverage: primary gauge is the incremental operating margin (ΔOpInc/ΔRev YoY) — robust when "
     "margins are negative; the Op Lev (x) ratio (OpInc growth ÷ Rev growth) is shown only when both are positive.", None),
    ("FCF: analyzed on an LTM basis (4-quarter sums) because quarterly FCF is lumpy.", None),
    ("Blanks: a blank metric means insufficient history (recent IPOs: FIG, RBRK, KVYO…), a non-positive base, or "
     "no estimate coverage — guarded on purpose rather than showing junk.", None),
    ("", None),
    ("WHAT WAS FIXED VS THE ORIGINAL QOQYOY SHEET", "H"),
    ("• Revenue-only → adds op income, FCF, margins, leverage and NTM revisions (the metrics you asked for).", None),
    ("• One 50-quarter spill per name (range(-40,10)) caused walls of #N/A for young names and 12,000+ stray "
     "columns → tightened to FPO −13…+2 with error-guarded metric formulas.", None),
    ("• Quarter headers were SAP's calendar applied to all rows → replaced with explicit FPO headers plus a "
     "per-company 'Latest Rpt Qtr' label pulled by BQL.", None),
    ("• Stale hardcoded labels ('2025 Q2 Estimate') and duplicate label rows → dynamic, per-company labels; "
     "mixed reporting currencies → USD-normalized pulls.", None),
    ("• Ticker universe = the two lists you supplied for this build (whitespace normalized), not the old sheet's "
     "list. Old-sheet names not on your new list were NOT carried over — several were acquired/delisted in 2025 "
     "(ALTR, ANSS, AZPN, BASE, INFA, SMAR, SWI…), but some still trade (CFLT, DBX, FFIV, JAMF, PRGS, QTWO, YEXT, "
     "YOU, COUR, OS; ZI renamed to GTM US Equity). To re-add any, insert them on the Tickers sheet and copy the "
     "sector-sheet formula rows down.", None),
    ("• No peer aggregation → dashboards with medians/quartiles/ranks and a Focus panel (ZM pre-loaded).", None),
    ("", None),
    ("COLOR LEGEND", "H"),
    ("Blue text = input you can edit (tickers, Focus company).  Black = formula.  Green = link to another sheet.  "
     "Yellow fill = the cell to type in.  Shaded columns (FPO +1/+2) = consensus estimates.  Grey italic = notes.", None),
    ("", None),
    ("KNOWN CAVEATS", "H"),
    ("• Verify unusual tickers on your terminal: 'P US Equity' (delisted Pandora?), '20 HK', '2241 CH' etc. — a "
     "row of #N/A means Bloomberg cannot resolve the ticker; fix it on the Tickers sheet.", None),
    ("• EPS revision % flips sign when EPS crosses zero — read with the revenue revision column.", None),
    ("• FCF estimates (FPO +1/+2) are sparse for smaller names; blanks there are normal.", None),
    ("• Sorting the Dashboard keeps each row internally consistent, but re-sorting is manual after a refresh.", None),
]
r = 1
for text, kind in rows:
    c = rd.cell(row=r, column=1, value=text)
    if kind == "T1":
        c.font = font(bold=True, size=15)
    elif kind == "sub":
        c.font = font(italic=True, color=NOTE_GREY)
    elif kind == "H":
        c.font = font(bold=True, size=11, color=WHITE)
        c.fill = HDR_FILL
    else:
        c.font = font()
    c.alignment = Alignment(wrap_text=True, vertical="top")
    rd.row_dimensions[r].height = None if kind in ("T1", "H") else max(
        15, 14 * (1 + len(text) // 110))
    r += 1
rd.column_dimensions["A"].width = 118

# sheet order: README, Tickers, then per sector Dashboard/Focus/Metrics/Revisions/Rev/OpInc/FCF
order = ["README", "Tickers"]
for code, *_ in SECTORS:
    order += [f"{code} Dashboard", f"{code} Focus", f"{code} Metrics",
              f"{code} Revisions", f"{code} Rev", f"{code} OpInc", f"{code} FCF"]
wb._sheets = [wb[s] for s in order]

wb.calculation.fullCalcOnLoad = True
wb.save(OUT)
print("saved", OUT)
import json
print(json.dumps({k[0] + ":" + k[1]: v[:2] for k, v in block_rows.items()}, indent=0))
