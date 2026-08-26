#!/usr/bin/env python3
"""Simulate a Bloomberg refresh with synthetic data and verify every native
Excel formula in the workbook computes what it should.

- Replaces all _xll.BQL spills / _xll.BDP calls with synthetic numbers
  (plus =NA() walls for young names, one dead-ticker row per sector,
  zero-revenue and zero-denominator cases to exercise every guard).
- Recalcs with LibreOffice, then compares every Metrics / Dashboard / Focus /
  Revisions formula cell against an independent Python reimplementation.
"""
import random, subprocess, sys, json
import openpyxl
from openpyxl.utils import get_column_letter

sys.path.insert(0, ".")
import build_workbook as B   # re-runs the build; gives us lists & layout consts

SRC = "Earnings_Sector_Comp_final.xlsx"
TST = "test_filled.xlsx"

SECTORS = [("SW", B.SOFTWARE, "ZM US Equity"), ("Semi", B.SEMIS, "NVDA US Equity")]
P_LO, P_HI = B.DATA_FPO_LO, B.DATA_FPO_HI
G_LO, G_HI = B.GRID_FPO_LO, B.GRID_FPO_HI
DR0 = B.DATA_START_ROW
RV0 = B.REVISION_ROW0
DASH0 = B.DASH_ROW0

rng = random.Random(42)

# ---------------------------------------------------------------- synth data
data = {}       # (code, kind)[i][p] -> float | None       kind in rev/opi/fcf
revisions = {}  # (code)[i] -> dict of 8 values (sales n/1/3/6, eps n/1/3/6), float|None
qlabels = {}

for code, lst, focus in SECTORS:
    n = len(lst)
    protected = {lst.index(focus)}
    full_na = {lst.index("P US Equity") if code == "SW" else lst.index("2241 CH Equity")}
    candidates = [i for i in range(n) if i not in protected | full_na]
    young = set(rng.sample(candidates, 12))
    fcf_est_na = set(rng.sample(candidates, 15))
    zero_rev_row = [i for i in candidates if i not in young][0]
    neg_opi_rows = set(rng.sample([i for i in candidates if i not in young], 8))
    rev, opi, fcf = {}, {}, {}
    for i in range(n):
        rev[i], opi[i], fcf[i] = {}, {}, {}
        base = rng.uniform(200, 8000)
        margin = rng.uniform(0.02, 0.35) if i not in neg_opi_rows else rng.uniform(-0.25, -0.02)
        g = rng.uniform(-0.02, 0.08)
        val = base
        k_na = rng.randint(4, 12) if i in young else 0
        for p in range(P_LO, P_HI + 1):
            val *= (1 + g + rng.gauss(0, 0.03))
            m = margin + (p - P_LO) * rng.uniform(0, 0.004) + rng.gauss(0, 0.02)
            rev[i][p] = round(val, 4)
            opi[i][p] = round(val * m, 4)
            fcf[i][p] = round(val * (m + rng.gauss(0, 0.05)), 4)
        if i in full_na:
            for p in range(P_LO, P_HI + 1):
                rev[i][p] = opi[i][p] = fcf[i][p] = None
        elif i in young:
            for p in range(P_LO, P_LO + k_na):
                rev[i][p] = opi[i][p] = fcf[i][p] = None
        if i in fcf_est_na and i not in full_na:
            fcf[i][1] = fcf[i][2] = None
        if i == zero_rev_row:
            rev[i][P_LO] = 0.0
    data[(code, "rev")], data[(code, "opi")], data[(code, "fcf")] = rev, opi, fcf
    qlabels[code] = {i: f"20{24 + i % 2} Q{1 + i % 4}" for i in range(n)}

    rv = {}
    zero_ago = set(rng.sample(candidates, 5))
    na_6m = set(rng.sample(candidates, 10))
    for i in range(n):
        base_s = rng.uniform(800, 40000)
        base_e = rng.uniform(-2, 12)
        d = {}
        for k, tag in enumerate(["now", "1m", "3m", "6m"]):
            d[f"s_{tag}"] = round(base_s * (1 + rng.gauss(0, 0.02) * k), 4)
            d[f"e_{tag}"] = round(base_e * (1 + rng.gauss(0, 0.05) * k), 4)
        if i in full_na:
            d = {k: None for k in d}
        if i in na_6m:
            d["s_6m"] = d["e_6m"] = None
        if i in zero_ago:
            d["s_1m"] = 0.0
            d["e_3m"] = 0.0
        rv[i] = d
    revisions[code] = rv

# ---------------------------------------------------------------- fill test wb
wb = openpyxl.load_workbook(SRC)
for code, lst, focus in SECTORS:
    n = len(lst)
    for suffix, kind in (("Rev", "rev"), ("OpInc", "opi"), ("FCF", "fcf")):
        ws = wb[f"{code} {suffix}"]
        for i in range(n):
            r = DR0 + i
            if suffix == "Rev":
                ws[f"B{r}"] = qlabels[code][i]
            for p in range(P_LO, P_HI + 1):
                v = data[(code, kind)][i][p]
                ws[f"{B.data_col(p)}{r}"] = "=NA()" if v is None else v
    ws = wb[f"{code} Revisions"]
    order = [("C", "s_now"), ("D", "s_1m"), ("E", "s_3m"), ("F", "s_6m"),
             ("K", "e_now"), ("L", "e_1m"), ("M", "e_3m"), ("N", "e_6m")]
    for i in range(n):
        r = RV0 + i
        for cl, key in order:
            v = revisions[code][i][key]
            ws[f"{cl}{r}"] = "=NA()" if v is None else v
wb.save(TST)
print("test workbook written; computing with the `formulas` evaluator…", flush=True)
# LibreOffice cannot load files in this environment, so evaluate with the
# `formulas` package instead (independent Excel-semantics implementation).
import re as _re
import formulas
from formulas.errors import BaseError
try:
    from formulas.functions import XlError
except ImportError:
    XlError = ()

xl_model = formulas.ExcelModel().loads(TST).finish()
solution = xl_model.calculate()
computed = {}
_pat = _re.compile(r"'\[[^\]]+\]([^']+)'!([A-Z]{1,3}[0-9]+)$")
for k, v in solution.items():
    m = _pat.match(k)
    if not m:
        continue
    sheet, coord = m.group(1).upper(), m.group(2)
    try:
        arr = v.value if hasattr(v, "value") else v
        val = arr[0][0] if getattr(arr, "ndim", None) == 2 else arr
    except Exception:
        val = None
    if val is not None and type(val).__name__ in ("str_", "float64", "int64", "bool_"):
        val = val.item()
    if XlError and isinstance(val, XlError):
        val = str(val)
    elif val is not None and not isinstance(val, (int, float, str, bool)):
        val = str(val)
    computed[(sheet, coord)] = val
print(f"computed {len(computed)} cells", flush=True)


class _Cell:
    def __init__(self, value):
        self.value = value


class _Sheet:
    def __init__(self, name):
        self._n = name.upper()

    def __getitem__(self, coord):
        return _Cell(computed.get((self._n, coord)))

# ---------------------------------------------------------------- expectations
BLANK = object()

def isnum(x):
    return isinstance(x, (int, float)) and x is not None

def expected_metric(code, key, i, p):
    rev = data[(code, "rev")][i]
    opi = data[(code, "opi")][i]
    fcf = data[(code, "fcf")][i]
    r0, r1, r4 = rev.get(p), rev.get(p - 1), rev.get(p - 4)
    o0, o4 = opi.get(p), opi.get(p - 4)
    if key == "rev":
        return r0 if isnum(r0) else BLANK
    if key == "rev_yoy":
        return r0 / r4 - 1 if isnum(r0) and isnum(r4) and r4 > 0 else BLANK
    if key == "rev_qoq":
        return r0 / r1 - 1 if isnum(r0) and isnum(r1) and r1 > 0 else BLANK
    if key == "add_yoy":
        return r0 - r4 if isnum(r0) and isnum(r4) else BLANK
    if key == "add_qoq":
        return r0 - r1 if isnum(r0) and isnum(r1) else BLANK
    if key == "opm":
        return o0 / r0 if isnum(o0) and isnum(r0) and r0 > 0 else BLANK
    if key == "opm_yoy":
        if all(isnum(x) for x in (o0, r0, o4, r4)) and r0 > 0 and r4 > 0:
            return (o0 / r0 - o4 / r4) * 10000
        return BLANK
    if key == "oi_yoy":
        return o0 / o4 - 1 if isnum(o0) and isnum(o4) and o4 > 0 else BLANK
    if key == "incr_opm":
        if all(isnum(x) for x in (o0, o4, r0, r4)) and abs(r0 - r4) >= 1:
            return (o0 - o4) / (r0 - r4)
        return BLANK
    if key == "fcf_ltm":
        rw = [rev.get(x) for x in range(p - 3, p + 1)]
        fw = [fcf.get(x) for x in range(p - 3, p + 1)]
        if any(not isnum(x) for x in rw + fw):
            return BLANK          # SUM over range with #N/A -> error -> IFERROR
        return BLANK if sum(rw) <= 0 else sum(fw) / sum(rw)
    if key == "fcf_ltm_yoy":
        if p - 7 < P_LO:
            return BLANK          # formula is '=""'
        now = expected_metric(code, "fcf_ltm", i, p)
        prior = expected_metric(code, "fcf_ltm", i, p - 4)
        if now is BLANK or prior is BLANK:
            return BLANK
        return (now - prior) * 10000
    raise KeyError(key)

def exp_revision(code, i, which, ago):
    d = revisions[code][i]
    cur, past = d[f"{which}_now"], d[f"{which}_{ago}"]
    if isnum(cur) and isnum(past) and past != 0:
        return cur / past - 1
    return BLANK

def exp_dash(code, lst, i):
    rev = data[(code, "rev")][i]
    e = {}
    e["C"] = rev[0] if isnum(rev[0]) else BLANK
    e["D"] = expected_metric(code, "rev_yoy", i, 0)
    e["E"] = expected_metric(code, "rev_yoy", i, -1)
    e["F"] = (e["D"] - e["E"]) if e["D"] is not BLANK and e["E"] is not BLANK else BLANK
    e["G"] = expected_metric(code, "rev_qoq", i, 0)
    e["H"] = expected_metric(code, "add_qoq", i, 0)
    e["I"] = expected_metric(code, "add_yoy", i, 0)
    e["J"] = expected_metric(code, "rev_yoy", i, 1)
    e["K"] = expected_metric(code, "opm", i, 0)
    e["L"] = expected_metric(code, "opm_yoy", i, 0)
    e["M"] = expected_metric(code, "oi_yoy", i, 0)
    e["N"] = expected_metric(code, "incr_opm", i, 0)
    d_, m_ = e["D"], e["M"]
    dd = d_ if d_ is not BLANK else 0
    mm = m_ if m_ is not BLANK else 0
    e["O"] = mm / dd if dd > 0 and mm > 0 else BLANK
    e["P"] = expected_metric(code, "fcf_ltm", i, 0)
    e["Q"] = expected_metric(code, "fcf_ltm_yoy", i, 0)
    e["R"] = exp_revision(code, i, "s", "1m")
    e["S"] = exp_revision(code, i, "s", "3m")
    e["T"] = exp_revision(code, i, "s", "6m")
    e["U"] = exp_revision(code, i, "e", "1m")
    e["V"] = exp_revision(code, i, "e", "3m")
    e["W"] = exp_revision(code, i, "e", "6m")
    return e

def percentile_inc(vals, q):
    vals = sorted(vals)
    if not vals:
        return BLANK
    k = (len(vals) - 1) * q
    f = int(k)
    c = min(f + 1, len(vals) - 1)
    return vals[f] + (vals[c] - vals[f]) * (k - f)

# ---------------------------------------------------------------- compare
out = {name: _Sheet(name) for name in wb.sheetnames}
fails = []
checked = 0

def isblank(v):
    return v is None or v == ""

def check(sheet, coord, exp, got):
    global checked
    checked += 1
    if exp is BLANK or exp is None:
        ok = isblank(got)
    elif isinstance(exp, str):
        ok = got == exp
    else:
        ok = isinstance(got, (int, float)) and abs(got - exp) <= max(1e-6, 1e-8 * abs(exp))
    if not ok:
        fails.append((sheet, coord, repr(exp)[:60], repr(got)[:60]))

for code, lst, focus in SECTORS:
    n = len(lst)
    met = out[f"{code} Metrics"]
    for key, *_ in B.BLOCKS:
        r0, r1, _ = B.block_rows[(code, key)]
        for i in range(n):
            for p in range(G_LO, G_HI + 1):
                exp = expected_metric(code, key, i, p)
                got = met[f"{B.grid_col(p)}{r0 + i}"].value
                check(f"{code} Metrics[{key}]", f"{B.grid_col(p)}{r0 + i}", exp, got)
    # revisions deltas
    rvs = out[f"{code} Revisions"]
    for i in range(n):
        r = RV0 + i
        for cl, which, ago in [("G", "s", "1m"), ("H", "s", "3m"), ("I", "s", "6m"),
                               ("O", "e", "1m"), ("P", "e", "3m"), ("Q", "e", "6m")]:
            check(f"{code} Revisions", f"{cl}{r}", exp_revision(code, i, which, ago),
                  rvs[f"{cl}{r}"].value)
    # dashboard rows
    dash = out[f"{code} Dashboard"]
    exps = [exp_dash(code, lst, i) for i in range(n)]
    for i in range(n):
        r = DASH0 + i
        check(f"{code} Dashboard", f"A{r}", lst[i], dash[f"A{r}"].value)
        check(f"{code} Dashboard", f"B{r}", qlabels[code][i], dash[f"B{r}"].value)
        for cl in "CDEFGHIJKLMNOPQRSTUVW":
            check(f"{code} Dashboard", f"{cl}{r}", exps[i][cl], dash[f"{cl}{r}"].value)
    # ranks
    dvals = [(i, e["D"]) for i, e in enumerate(exps) if e["D"] is not BLANK]
    svals = [(i, e["S"]) for i, e in enumerate(exps) if e["S"] is not BLANK]
    for i in range(n):
        r = DASH0 + i
        for cl, pool, e in (("X", dvals, exps[i]["D"]), ("Y", svals, exps[i]["S"])):
            exp = BLANK if e is BLANK else 1 + sum(1 for _, v in pool if v > e)
            check(f"{code} Dashboard", f"{cl}{r}", exp, dash[f"{cl}{r}"].value)
    # dashboard stat rows
    for k, (lab, q_) in enumerate([("med", None), ("q3", 0.75), ("q1", 0.25)]):
        sr = B.DASH_ROW_HDR + 1 + k
        for cl in "CDEFGHIJKLMNOPQRSTUVW":
            vals = [e[cl] for e in exps if e[cl] is not BLANK]
            if not vals:
                exp = BLANK
            elif q_ is None:
                exp = percentile_inc(vals, 0.5)
            else:
                exp = percentile_inc(vals, q_)
            check(f"{code} Dashboard stats", f"{cl}{sr}", exp, dash[f"{cl}{sr}"].value)
    # focus sheet
    foc = out[f"{code} Focus"]
    fi = lst.index(focus)
    check(f"{code} Focus", "C4", f"Latest reported qtr: {qlabels[code][fi]}", foc["C4"].value)
    r = 7
    for name, key, fmt, kind in B.FOCUS_METRICS:
        for series in ("Company", "Peer median", "vs median"):
            for p in range(G_LO, G_HI + 1):
                cl = B.grid_col(p)
                comp = expected_metric(code, key, fi, p)
                pool = [expected_metric(code, key, i, p) for i in range(n)]
                pool = [v for v in pool if v is not BLANK and v is not None]
                medv = percentile_inc(pool, 0.5) if pool else BLANK
                if series == "Company":
                    exp = comp
                elif series == "Peer median":
                    exp = medv
                else:
                    exp = comp - medv if comp not in (BLANK, None) and medv is not BLANK else BLANK
                check(f"{code} Focus[{key}:{series}]", f"{cl}{r}", exp, foc[f"{cl}{r}"].value)
            r += 1
        r += 1
    # focus revisions mini-block
    r += 2  # block title + header
    for which in ("s", "e"):
        for series in ("Company", "Peer median", "vs median"):
            for k, ago in enumerate(("1m", "3m", "6m")):
                cl = "CDE"[k]
                comp = exp_revision(code, fi, which, ago)
                pool = [exp_revision(code, i, which, ago) for i in range(n)]
                pool = [v for v in pool if v is not BLANK]
                medv = percentile_inc(pool, 0.5) if pool else BLANK
                if series == "Company":
                    exp = comp
                elif series == "Peer median":
                    exp = medv
                else:
                    exp = comp - medv if comp is not BLANK and medv is not BLANK else BLANK
                check(f"{code} Focus[rev {which}:{series}]", f"{cl}{r}", exp, foc[f"{cl}{r}"].value)
            r += 1
        r += 1

# ---------------------------------------------------------------- error audit
err_cells = {}
sheet_by_upper = {name.upper(): name for name in wb.sheetnames}
for (sheet_u, coord), v in computed.items():
    if isinstance(v, str) and v.startswith("#"):
        err_cells.setdefault(sheet_by_upper.get(sheet_u, sheet_u), []).append((coord, v))
allowed = {f"{code} {s}" for code, *_ in SECTORS for s in ("Rev", "OpInc", "FCF", "Revisions")}
bad_err = {k: v for k, v in err_cells.items() if k not in allowed}
n_inj = sum(len(v) for k, v in err_cells.items() if k in allowed)

print(f"\nchecked {checked} cells; mismatches: {len(fails)}")
for f_ in fails[:60]:
    print("  MISMATCH", f_)
print(f"injected #N/A cells on data sheets: {n_inj}")
print("errors OUTSIDE data sheets:", json.dumps(bad_err)[:2000] or "{}")
if fails or bad_err:
    sys.exit(1)
print("ALL FORMULA CHECKS PASSED")
